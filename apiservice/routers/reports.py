import os
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Optional

import psycopg2.extras
from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException
from fastapi.responses import FileResponse

from db import get_conn, get_db
from schemas.reports import ReportCreate

router = APIRouter()

REPORTS_DIR = Path(os.environ.get("REPORTS_DIR", "/tmp/reports"))


# ── Background report generation ─────────────────────────────────────────────

def _fetch_summary(conn, period_start: date, period_end: date, zone_id: Optional[int]) -> dict:
    params = {"from_dt": period_start, "to_dt": period_end}
    z_cnt = "AND cnt.zone_id = %(zone_id)s" if zone_id else ""
    z_ads = "AND zone_id   = %(zone_id)s" if zone_id else ""
    if zone_id:
        params["zone_id"] = zone_id

    with conn.cursor() as cur:
        cur.execute("SELECT COUNT(*) FROM containers WHERE is_active = true")
        active = cur.fetchone()[0]

    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute(f"""
            SELECT COUNT(*) AS cnt, COALESCE(SUM(c.volume_collected_l), 0) AS vol
            FROM collections c
            JOIN containers cnt ON cnt.key_container = c.container_id
            WHERE c.collected_at::date BETWEEN %(from_dt)s AND %(to_dt)s {z_cnt}
        """, params)
        col = cur.fetchone()

    with conn.cursor() as cur:
        cur.execute(f"""
            SELECT ROUND(AVG(avg_fill_rate)::numeric, 2),
                   COALESCE(SUM(overflow_count), 0)
            FROM aggregated_daily_stats
            WHERE day_bucket BETWEEN %(from_dt)s AND %(to_dt)s {z_ads}
        """, params)
        r = cur.fetchone()
        avg_fill = float(r[0] or 0)
        overflow = int(r[1])

    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        z_zone = "AND z.key_zone = %(zone_id)s" if zone_id else ""
        cur.execute(f"""
            SELECT z.name, COUNT(c.key_collection) AS cnt,
                   COALESCE(SUM(c.volume_collected_l), 0) AS vol
            FROM zones z
            JOIN containers cnt ON cnt.zone_id = z.key_zone
            JOIN collections c  ON c.container_id = cnt.key_container
            WHERE c.collected_at::date BETWEEN %(from_dt)s AND %(to_dt)s {z_zone}
            GROUP BY z.name ORDER BY vol DESC LIMIT 10
        """, params)
        zones = cur.fetchall()

    return {
        "active_containers": active,
        "collection_count":  int(col["cnt"]),
        "volume_l":          float(col["vol"]),
        "avg_fill_rate":     avg_fill,
        "overflow_count":    overflow,
        "top_zones":         [(r["name"] or "—", r["cnt"], float(r["vol"])) for r in zones],
    }


def _build_pdf(report_id: int, report_type: str, summary: dict,
               period_start: date, period_end: date, file_path: Path) -> None:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    doc = SimpleDocTemplate(str(file_path), pagesize=A4)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph(f"ECOTRACK — {report_type.capitalize()} Report", styles["Title"]))
    story.append(Paragraph(f"Period: {period_start} → {period_end}", styles["Normal"]))
    story.append(Paragraph(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", styles["Normal"]))
    story.append(Spacer(1, 16))

    kpi_data = [
        ["KPI", "Value"],
        ["Active containers",  str(summary["active_containers"])],
        ["Collections",        str(summary["collection_count"])],
        ["Volume collected (L)", f"{summary['volume_l']:,.0f}"],
        ["Avg fill rate",      f"{summary['avg_fill_rate']:.1f} %"],
        ["Overflow events",    str(summary["overflow_count"])],
    ]
    t = Table(kpi_data, hAlign="LEFT", colWidths=[220, 160])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2d6a4f")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTSIZE",   (0, 0), (-1, -1), 10),
        ("GRID",       (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f0f4f0")]),
    ]))
    story.append(Paragraph("Key Performance Indicators", styles["Heading2"]))
    story.append(t)
    story.append(Spacer(1, 16))

    if summary["top_zones"]:
        story.append(Paragraph("Top Zones by Volume Collected", styles["Heading2"]))
        zone_data = [["Zone", "Collections", "Volume (L)"]]
        for name, cnt, vol in summary["top_zones"]:
            zone_data.append([name, str(cnt), f"{vol:,.0f}"])
        zt = Table(zone_data, hAlign="LEFT", colWidths=[220, 100, 120])
        zt.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1b4332")),
            ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
            ("FONTSIZE",   (0, 0), (-1, -1), 9),
            ("GRID",       (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f0f4f0")]),
        ]))
        story.append(zt)

    doc.build(story)


def _build_excel(report_id: int, report_type: str, summary: dict,
                 period_start: date, period_end: date, file_path: Path) -> None:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()

    # Sheet 1 — KPIs
    ws1 = wb.active
    ws1.title = "Summary"
    header_fill = PatternFill("solid", fgColor="2D6A4F")
    header_font = Font(bold=True, color="FFFFFF")

    ws1.append([f"ECOTRACK — {report_type.capitalize()} Report"])
    ws1.append([f"Period: {period_start} → {period_end}"])
    ws1.append([])
    ws1.append(["KPI", "Value"])
    for cell in ws1[4]:
        cell.fill = header_fill
        cell.font = header_font
    ws1.append(["Active containers",     summary["active_containers"]])
    ws1.append(["Collections",           summary["collection_count"]])
    ws1.append(["Volume collected (L)",  summary["volume_l"]])
    ws1.append(["Avg fill rate (%)",     summary["avg_fill_rate"]])
    ws1.append(["Overflow events",       summary["overflow_count"]])
    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 16

    # Sheet 2 — Zone breakdown
    ws2 = wb.create_sheet("Zones")
    ws2.append(["Zone", "Collections", "Volume (L)"])
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
    for name, cnt, vol in summary["top_zones"]:
        ws2.append([name, cnt, vol])
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 16

    wb.save(str(file_path))


def _generate_report(report_id: int, report_type: str, fmt: str,
                     period_start: date, period_end: date,
                     zone_id: Optional[int]) -> None:
    """Background task — generates the file and updates the reports table."""
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    ext = "xlsx" if fmt == "excel" else "pdf"
    file_path = REPORTS_DIR / f"{report_id}.{ext}"

    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "UPDATE reports SET status = 'processing' WHERE key_report = %s",
                (report_id,),
            )

    try:
        with get_conn() as conn:
            summary = _fetch_summary(conn, period_start, period_end, zone_id)

        if fmt == "excel":
            _build_excel(report_id, report_type, summary, period_start, period_end, file_path)
        else:
            _build_pdf(report_id, report_type, summary, period_start, period_end, file_path)

        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "UPDATE reports SET status = 'ready', file_path = %s WHERE key_report = %s",
                    (str(file_path), report_id),
                )
    except Exception:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "UPDATE reports SET status = 'error' WHERE key_report = %s",
                    (report_id,),
                )


# ── Endpoints ─────────────────────────────────────────────────────────────────

@router.post("/generate", status_code=202)
def generate_report(
    body: ReportCreate,
    background_tasks: BackgroundTasks,
    conn=Depends(get_db),
):
    """
    Trigger async report generation. Returns immediately with report_id.
    Poll GET /reports/{id}/download — streams file when status = 'ready'.
    """
    period_start = body.period_start or (date.today().replace(day=1) if body.report_type == "monthly"
                                         else date.today() - timedelta(days=7))
    period_end   = body.period_end   or date.today()

    with conn.cursor() as cur:
        cur.execute(
            """
            INSERT INTO reports (user_id, zone_id, report_type, format, period_start, period_end, status)
            VALUES (%s, %s, %s, %s, %s, %s, 'pending')
            RETURNING key_report
            """,
            (body.user_id, body.zone_id, body.report_type, body.format, period_start, period_end),
        )
        report_id = cur.fetchone()[0]

    background_tasks.add_task(
        _generate_report,
        report_id, body.report_type, body.format,
        period_start, period_end, body.zone_id,
    )
    return {"report_id": report_id, "status": "pending"}


@router.get("/{report_id}/download")
def download_report(report_id: int, conn=Depends(get_db)):
    """Stream the generated file. Returns 202 while still processing, 404 if unknown."""
    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute(
            "SELECT status, file_path, format, report_type, period_start FROM reports WHERE key_report = %s",
            (report_id,),
        )
        row = cur.fetchone()

    if not row:
        raise HTTPException(status_code=404, detail="Report not found")
    if row["status"] == "pending":
        return {"status": "pending"}
    if row["status"] == "processing":
        return {"status": "processing"}
    if row["status"] == "error":
        raise HTTPException(status_code=500, detail="Report generation failed")

    file_path = Path(row["file_path"])
    if not file_path.exists():
        raise HTTPException(status_code=500, detail="Report file missing from storage")

    if row["format"] == "excel":
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = f"report_{row['report_type']}_{row['period_start']}.xlsx"
    else:
        media_type = "application/pdf"
        filename = f"report_{row['report_type']}_{row['period_start']}.pdf"

    return FileResponse(
        path=str(file_path),
        media_type=media_type,
        filename=filename,
    )
